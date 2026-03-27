"""
EduTrack — Excel to Firebase Migration Script
=============================================
Transfers your Google Sheet / Excel data directly into Firebase Firestore.

SETUP:
1. pip install firebase-admin openpyxl gspread google-auth pandas
2. Download your Firebase Service Account key:
   Firebase Console → Project Settings → Service accounts → Generate new private key
   Save as serviceAccountKey.json in this folder
3. Export your Google Sheet as Excel (.xlsx) OR provide the Google Sheet ID
4. Run: python migrate_to_firebase.py

The script will:
  - Read all lesson records from your Excel / Google Sheet
  - Create Users in Firebase Auth (teacher, students, parent)
  - Upload all lesson logs to Firestore 'lessonLogs' collection
  - Print a summary when done
"""

import json
import sys
from datetime import datetime

# ── CONFIG ──────────────────────────────────────────
SERVICE_ACCOUNT_KEY = 'serviceAccountKey.json'  # Download from Firebase Console

# Option A: Use Excel file (export your Google Sheet as .xlsx)
EXCEL_FILE = 'Academic_Progress_Report.xlsx'

# Option B: Use Google Sheets directly (provide sheet ID)
# SHEET_ID = '1gRX9_ozV6cfdxKXsbSUS6M3_sUMNFAgb8v7FJqK_o2c'

# Student info (update these!)
STUDENT_NAME = 'Sahil'
STUDENT_CLASS = 'Form 4'
CURRICULUM = 'KSSM'

# Default user accounts to create
USERS = [
    {'email': 'teacher@school.edu.my', 'password': 'teacher123', 'role': 'teacher', 'name': 'Admin Teacher', 'linkedStudent': '', 'class': ''},
    {'email': 'sahil@student.edu.my',  'password': 'student123', 'role': 'student', 'name': 'Sahil',          'linkedStudent': 'Sahil', 'class': 'Form 4'},
    {'email': 'parent@sahil.com',      'password': 'parent123',  'role': 'parent',  'name': "Sahil's Parent", 'linkedStudent': 'Sahil', 'class': 'Form 4'},
]
# ────────────────────────────────────────────────────

try:
    import firebase_admin
    from firebase_admin import credentials, firestore, auth
    import openpyxl
    import pandas as pd
except ImportError:
    print("❌ Missing packages. Run: pip install firebase-admin openpyxl pandas")
    sys.exit(1)


def init_firebase():
    try:
        cred = credentials.Certificate(SERVICE_ACCOUNT_KEY)
        firebase_admin.initialize_app(cred)
        print("✅ Firebase connected")
        return firestore.client()
    except Exception as e:
        print(f"❌ Firebase init failed: {e}")
        print("  → Make sure serviceAccountKey.json is in this folder")
        sys.exit(1)


def parse_excel(filepath):
    """Parse the Excel file in your current format (Sahil's IR Report)"""
    logs = []
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        print(f"📂 Found sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Read student metadata (rows 3-5)
            student_name = ''
            curriculum = CURRICULUM
            level = STUDENT_CLASS
            try:
                student_name = str(ws.cell(3, 3).value or '').strip() or STUDENT_NAME
                curriculum   = str(ws.cell(4, 3).value or CURRICULUM).strip()
                level        = str(ws.cell(5, 3).value or STUDENT_CLASS).strip()
            except:
                student_name = STUDENT_NAME

            if not student_name:
                student_name = STUDENT_NAME

            print(f"\n📋 Sheet: {sheet_name} → Student: {student_name}")

            # Data starts at row 9 (index 8), headers at row 8
            headers_row = 8
            data_start  = 9

            # Read each data row
            for row_num in range(data_start, ws.max_row + 1):
                date    = str(ws.cell(row_num, 1).value or '').strip()
                subj_raw= str(ws.cell(row_num, 2).value or '').strip()
                topic   = str(ws.cell(row_num, 3).value or '').strip()
                progress= str(ws.cell(row_num, 4).value or '').strip()
                homework= str(ws.cell(row_num, 5).value or '').strip()
                obs     = str(ws.cell(row_num, 6).value or '').strip()
                followup= str(ws.cell(row_num, 7).value or '').strip()

                # Skip empty rows
                if not date and not topic:
                    continue
                # Skip exam rows
                if 'examination' in date.lower() or 'term' in date.lower():
                    continue
                # Skip rows with no real content
                if not subj_raw and not topic:
                    continue

                # Parse "Sejarah/Ms Amirah" → subject + teacher
                subject = subj_raw
                teacher = ''
                if '/' in subj_raw:
                    parts   = subj_raw.split('/', 1)
                    subject = parts[0].strip()
                    teacher = parts[1].strip()

                # Skip broken references
                if subject.startswith('#') or topic.startswith('#'):
                    continue

                log = {
                    'studentName': student_name,
                    'date':        date,
                    'subject':     subject,
                    'teacher':     teacher,
                    'topic':       topic,
                    'progress':    progress,
                    'homework':    homework,
                    'observations':obs,
                    'followUp':    followup,
                    'class':       level,
                    'curriculum':  curriculum,
                    'createdAt':   datetime.now(),
                    'source':      'migrated_from_excel',
                }
                logs.append(log)
                print(f"  ✓ Row {row_num}: {date} | {subject} | {topic[:40]}...")

        print(f"\n✅ Parsed {len(logs)} lesson records")
        return logs

    except FileNotFoundError:
        print(f"❌ File not found: {filepath}")
        print("  → Export your Google Sheet as .xlsx and place it in this folder")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")
        raise


def upload_logs(db, logs):
    """Upload lesson logs to Firestore"""
    print(f"\n📤 Uploading {len(logs)} lesson records to Firestore...")
    collection = db.collection('lessonLogs')

    # Upload in batches of 500
    batch_size = 500
    for i in range(0, len(logs), batch_size):
        batch = db.batch()
        chunk = logs[i:i + batch_size]
        for log in chunk:
            ref = collection.document()
            batch.set(ref, log)
        batch.commit()
        print(f"  ✅ Uploaded batch {i // batch_size + 1} ({len(chunk)} records)")

    print(f"✅ All {len(logs)} records uploaded to 'lessonLogs' collection")


def create_users(db):
    """Create Firebase Auth users and Firestore user profiles"""
    print("\n👤 Creating user accounts...")
    for u in USERS:
        try:
            # Create Firebase Auth user
            try:
                user = auth.create_user(email=u['email'], password=u['password'], display_name=u['name'])
                uid = user.uid
                print(f"  ✅ Auth: {u['email']} (uid: {uid})")
            except auth.EmailAlreadyExistsError:
                user = auth.get_user_by_email(u['email'])
                uid = user.uid
                print(f"  ℹ️  Auth: {u['email']} already exists (uid: {uid})")

            # Save user profile to Firestore
            db.collection('users').document(uid).set({
                'email':         u['email'],
                'name':          u['name'],
                'role':          u['role'],
                'linkedStudent': u['linkedStudent'],
                'class':         u['class'],
                'createdAt':     datetime.now(),
            })
            print(f"  ✅ Firestore profile saved: {u['name']} ({u['role']})")

        except Exception as e:
            print(f"  ❌ Error creating {u['email']}: {e}")


def set_firestore_rules():
    print("""
╔══════════════════════════════════════════════════════════════╗
║  IMPORTANT: Set these Firestore Security Rules in Firebase   ║
╚══════════════════════════════════════════════════════════════╝

Go to Firebase Console → Firestore → Rules → paste this:

rules_version = '2';
service cloud.firestore {
  match /databases/{database}/documents {
    // Users can read their own profile
    match /users/{userId} {
      allow read, write: if request.auth.uid == userId;
      allow read: if request.auth != null
        && get(/databases/$(database)/documents/users/$(request.auth.uid)).data.role == 'teacher';
    }

    // Lesson logs
    match /lessonLogs/{docId} {
      allow read: if request.auth != null && (
        // Teachers can read all
        get(/databases/$(database)/documents/users/$(request.auth.uid)).data.role == 'teacher' ||
        // Students/parents can only read their linked student's logs
        resource.data.studentName ==
          get(/databases/$(database)/documents/users/$(request.auth.uid)).data.linkedStudent
      );
      allow create: if request.auth != null
        && get(/databases/$(database)/documents/users/$(request.auth.uid)).data.role == 'teacher';
    }

    // Exam results — same rules as lesson logs
    match /examResults/{docId} {
      allow read: if request.auth != null && (
        get(/databases/$(database)/documents/users/$(request.auth.uid)).data.role == 'teacher' ||
        resource.data.studentName ==
          get(/databases/$(database)/documents/users/$(request.auth.uid)).data.linkedStudent
      );
      allow write: if request.auth != null
        && get(/databases/$(database)/documents/users/$(request.auth.uid)).data.role == 'teacher';
    }
  }
}
""")


def main():
    print("=" * 60)
    print("  EduTrack — Excel to Firebase Migration")
    print("=" * 60)

    # 1. Connect to Firebase
    db = init_firebase()

    # 2. Parse Excel
    logs = parse_excel(EXCEL_FILE)

    if not logs:
        print("❌ No lesson records found. Check the Excel file format.")
        sys.exit(1)

    # 3. Confirm upload
    print(f"\n📊 Ready to upload:")
    print(f"   • {len(logs)} lesson records")
    print(f"   • {len(USERS)} user accounts")

    confirm = input("\nProceed? (yes/no): ").strip().lower()
    if confirm != 'yes':
        print("Cancelled.")
        sys.exit(0)

    # 4. Upload
    upload_logs(db, logs)
    create_users(db)
    set_firestore_rules()

    print("\n" + "=" * 60)
    print("  ✅ Migration Complete!")
    print("=" * 60)
    print(f"\n  {len(logs)} lesson records → Firestore 'lessonLogs'")
    print(f"  {len(USERS)} users → Firebase Auth + Firestore 'users'")
    print("\n  Next steps:")
    print("  1. Set the Firestore Security Rules shown above")
    print("  2. Run 'flutterfire configure' in your Flutter project")
    print("  3. Replace the values in lib/firebase_options.dart")
    print("  4. Run 'flutter run' to launch the app")
    print("\n  Login with:")
    for u in USERS:
        print(f"  {u['role'].capitalize()}: {u['email']} / {u['password']}")


if __name__ == '__main__':
    main()
